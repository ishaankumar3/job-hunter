#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Job Application Tracker — Ishaan Kumar
- Logs every job sent in the daily email
- Lets Ishaan mark Applied / Not Applied
- "Why No?" column lets Ishaan explain rejections
- Bot reads those reasons and avoids similar jobs in future
- Outcome tracking: Interview / Offer / Rejected
"""

import os
import re
import hashlib
import logging
from datetime import datetime

log = logging.getLogger(__name__)

TRACKER_FILE = "job_tracker.xlsx"

# A=hidden ID, B-J=bot fills, K-O=Ishaan fills
HEADERS = [
    "Job ID",           # A - hidden
    "Date Sent",        # B
    "Job Title",        # C
    "Company",          # D
    "Location",         # E
    "Salary",           # F
    "Source",           # G
    "Sponsorship",      # H
    "Score",            # I
    "Job URL",          # J
    "Applied?",         # K  YES / NO  (user)
    "Date Applied",     # L  (user)
    "Outcome",          # M  (user)
    "Why No?",          # N  (user — reason for NOT applying)
    "Notes",            # O  (user — free text)
]

COL_WIDTHS = {
    "A": 14, "B": 13, "C": 32, "D": 24,
    "E": 20, "F": 20, "G": 13, "H": 18,
    "I":  7, "J": 45, "K": 11, "L": 14,
    "M": 20, "N": 40, "O": 30,
}

# Pre-defined "Why No?" options for the dropdown
WHY_NO_OPTIONS = [
    "Too junior / entry level",
    "Too senior / overqualified",
    "Wrong sector / not water",
    "Wrong location / too far",
    "Salary too low",
    "Contract / not permanent",
    "Missing required licence/cert",
    "Company culture / reputation",
    "Role not interesting",
    "Duplicate / already applied",
    "Other",
]


def _normalize_company(company):
    """Normalize company name to reduce false duplicates."""
    c = company.lower().strip()
    for suffix in [" ltd", " limited", " plc", " llp", " uk", " consulting", 
                   " consultancy", " group", " holdings", " solutions", " services"]:
        if c.endswith(suffix):
            c = c[:-len(suffix)].strip()
    c = c.replace(".", "").replace(",", "").replace("&", "and")
    c = " ".join(c.split())
    return c

def _normalize_title(title):
    """Normalize title to catch near-duplicates."""
    t = title.lower().strip()
    t = t.replace("(", " ").replace(")", " ").replace("/", " ").replace("-", " ")
    t = " ".join(t.split())
    return t

def _job_id(title, company, url):
    """Generate stable job ID that resists company name variations."""
    norm_title = _normalize_title(title)
    norm_company = _normalize_company(company)
    
    # If URL contains unique ID, use that
    url_id = re.search(r'/jobs?/view/(\d+)|job[_-]?id[=:](\d+)|(\d{6,})', url)
    if url_id:
        unique = url_id.group(1) or url_id.group(2) or url_id.group(3)
        raw = unique + "|" + norm_company
    else:
        raw = norm_title + "|" + norm_company
    
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def _get_wb():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    if os.path.exists(TRACKER_FILE):
        try:
            wb = load_workbook(TRACKER_FILE)
            log.info("[TRACKER] Loaded: %s", TRACKER_FILE)
            return wb
        except Exception as e:
            log.warning("[TRACKER] Reload failed (%s) — creating fresh", e)

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Job Applications"

    bot_fill  = PatternFill("solid", fgColor="1A3A5C")
    user_fill = PatternFill("solid", fgColor="1A6FA8")
    why_fill  = PatternFill("solid", fgColor="744210")   # amber — stands out
    bot_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    user_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Border(
                    left  =Side(style="thin", color="CCCCCC"),
                    right =Side(style="thin", color="CCCCCC"),
                    top   =Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"),
                )

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = center
        cell.border    = thin
        if col_idx <= 10:                  # bot columns — navy
            cell.fill = bot_fill
            cell.font = bot_font
        elif header == "Why No?":          # special amber header
            cell.fill = why_fill
            cell.font = user_font
        else:                              # other user columns — blue
            cell.fill = user_fill
            cell.font = user_font

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].hidden = True

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Dropdowns
    dv_applied = DataValidation(type="list", formula1='"Yes,No"',
                                showDropDown=False, sqref="K2:K5000")
    dv_outcome = DataValidation(
        type="list",
        formula1='"Pending,Interview Invited,2nd Interview,Offer Received,Rejected,Withdrawn"',
        showDropDown=False, sqref="M2:M5000")
    why_no_str = '"' + ",".join(WHY_NO_OPTIONS) + '"'
    dv_why = DataValidation(type="list", formula1=why_no_str,
                            showDropDown=False, sqref="N2:N5000")
    ws.add_data_validation(dv_applied)
    ws.add_data_validation(dv_outcome)
    ws.add_data_validation(dv_why)

    # ── Stats sheet ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Stats")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16

    title_cell = ws2["A1"]
    title_cell.value = "Application Stats — Ishaan Kumar"
    title_cell.font  = Font(name="Arial", bold=True, size=13, color="1A3A5C")
    ws2.merge_cells("A1:B1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    stat_fill   = PatternFill("solid", fgColor="F0F4F8")
    s_thin      = Border(left =Side(style="thin", color="DDDDDD"),
                         right=Side(style="thin", color="DDDDDD"),
                         top  =Side(style="thin", color="DDDDDD"),
                         bottom=Side(style="thin", color="DDDDDD"))
    stats = [
        ("Total Jobs Sent",            "=COUNTA('Job Applications'!B2:B5000)"),
        ("Applied",                    "=COUNTIF('Job Applications'!K2:K5000,\"Yes\")"),
        ("Not Applied (skipped)",      "=COUNTIF('Job Applications'!K2:K5000,\"No\")"),
        ("Awaiting Decision",          "=COUNTIF('Job Applications'!M2:M5000,\"Pending\")"),
        ("Interviews",                 "=COUNTIF('Job Applications'!M2:M5000,\"Interview Invited\")"),
        ("2nd Interviews",             "=COUNTIF('Job Applications'!M2:M5000,\"2nd Interview\")"),
        ("Offers",                     "=COUNTIF('Job Applications'!M2:M5000,\"Offer Received\")"),
        ("Rejected by employer",       "=COUNTIF('Job Applications'!M2:M5000,\"Rejected\")"),
        ("Application Rate",           "=IFERROR(TEXT(B3/B2,\"0%\"),\"-\")"),
        ("Interview Rate",             "=IFERROR(TEXT(B6/B3,\"0%\"),\"-\")"),
        ("---", "---"),
        ("Why No — Too junior",        "=COUNTIF('Job Applications'!N2:N5000,\"Too junior / entry level\")"),
        ("Why No — Too senior",        "=COUNTIF('Job Applications'!N2:N5000,\"Too senior / overqualified\")"),
        ("Why No — Wrong sector",      "=COUNTIF('Job Applications'!N2:N5000,\"Wrong sector / not water\")"),
        ("Why No — Wrong location",    "=COUNTIF('Job Applications'!N2:N5000,\"Wrong location / too far\")"),
        ("Why No — Salary too low",    "=COUNTIF('Job Applications'!N2:N5000,\"Salary too low\")"),
        ("Why No — Contract role",     "=COUNTIF('Job Applications'!N2:N5000,\"Contract / not permanent\")"),
        ("Why No — Missing cert",      "=COUNTIF('Job Applications'!N2:N5000,\"Missing required licence/cert\")"),
        ("Why No — Other",             "=COUNTIF('Job Applications'!N2:N5000,\"Other\")"),
    ]

    for i, (label, formula) in enumerate(stats, 3):
        lc = ws2.cell(row=i, column=1, value=label)
        vc = ws2.cell(row=i, column=2, value=formula if formula != "---" else "")
        lc.font = Font(name="Arial", size=10, bold=(label != "---"))
        vc.font = Font(name="Arial", size=10)
        vc.alignment = Alignment(horizontal="center")
        if label != "---":
            lc.fill = stat_fill
            vc.fill = stat_fill
            lc.border = s_thin
            vc.border = s_thin

    log.info("[TRACKER] Created fresh tracker")
    return wb


# ================================================================
#  PUBLIC FUNCTIONS
# ================================================================

def load_sent_ids():
    """Return set of job IDs already sent."""
    if not os.path.exists(TRACKER_FILE):
        return set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(TRACKER_FILE, read_only=True)
        ws = wb.active
        ids = {str(row[0]) for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]}
        wb.close()
        log.info("[TRACKER] %d jobs already sent", len(ids))
        return ids
    except Exception as e:
        log.warning("[TRACKER] load_sent_ids failed: %s", e)
        return set()


def load_rejection_reasons():
    """
    Read tracker rows where Applied?=No and return a structured list of reasons.
    Used by job_search.py to filter/penalise similar future jobs.

    Returns list of dicts:
      { "title": str, "company": str, "reason": str, "reason_category": str }
    """
    if not os.path.exists(TRACKER_FILE):
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(TRACKER_FILE, read_only=True)
        ws = wb.active

        rejections = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            applied  = str(row[10] or "").strip().lower() if len(row) > 10 else ""
            why_no   = str(row[13] or "").strip()         if len(row) > 13 else ""
            title    = str(row[2]  or "").strip()         if len(row) > 2  else ""
            company  = str(row[3]  or "").strip()         if len(row) > 3  else ""

            if applied == "no":
                rejections.append({
                    "title":    title,
                    "company":  company,
                    "reason":   why_no,
                })

        wb.close()
        log.info("[TRACKER] %d rejection reasons loaded", len(rejections))
        return rejections
    except Exception as e:
        log.warning("[TRACKER] load_rejection_reasons failed: %s", e)
        return []


def should_skip_job(job, rejections):
    """
    Check whether a job should be filtered out based on past rejection reasons.
    Returns (skip: bool, reason: str)
    """
    if not rejections:
        return False, ""

    title   = job.get("title", "").lower()
    company = job.get("company", "").lower()
    salary_min = job.get("salary_min", 0) or 0

    for r in rejections:
        cat = r.get("reason", "")
        r_company = r.get("company", "").lower()
        r_title   = r.get("title",   "").lower()

        # Same company + same role type = skip
        if r_company and r_company in company and r_title and _title_similar(r_title, title):
            return True, "Similar to previously rejected: " + r["title"] + " @ " + r["company"]

        # Category-based rules
        if cat == "Too junior / entry level":
            if any(w in title for w in ["graduate", "junior", "trainee", "apprentice", "entry"]):
                return True, "Filtered: too junior (your feedback)"

        elif cat == "Too senior / overqualified":
            if any(w in title for w in ["director", "head of", "principal", "chief", "vp ", "vice president"]):
                return True, "Filtered: too senior (your feedback)"

        elif cat == "Wrong sector / not water":
            # Only skip non-water jobs if this reason appears 2+ times
            count = sum(1 for x in rejections if x.get("reason") == cat)
            if count >= 2:
                water_words = ["water", "hydraulic", "wastewater", "drainage", "sewer",
                               "pipeline", "nav", "utilities", "flood"]
                if not any(w in title for w in water_words):
                    return True, "Filtered: wrong sector (your feedback)"

        elif cat == "Wrong location / too far":
            pass  # location filtering needs more context — skip for now

        elif cat == "Salary too low":
            count = sum(1 for x in rejections if x.get("reason") == cat)
            if count >= 3 and salary_min and salary_min < 35000:
                return True, "Filtered: likely below your salary threshold (your feedback)"

        elif cat == "Contract / not permanent":
            if any(w in title.lower() for w in ["contract", "interim", "temp ", "temporary", "freelance"]):
                return True, "Filtered: contract role (your feedback)"

    return False, ""


def _title_similar(t1, t2):
    """Rough similarity — share 2+ meaningful words."""
    skip = {"engineer", "engineering", "senior", "junior", "the", "and", "for", "in", "at"}
    w1 = {w.strip("(),.-") for w in t1.split() if len(w) > 3 and w not in skip}
    w2 = {w.strip("(),.-") for w in t2.split() if len(w) > 3 and w not in skip}
    return len(w1 & w2) >= 2


def append_jobs_to_tracker(jobs):
    """Append new jobs to tracker. Returns (added_list, skipped_count)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    existing_ids = load_sent_ids()
    wb  = _get_wb()
    ws  = wb.active

    next_row = ws.max_row + 1
    if next_row == 2 and ws.cell(2, 1).value is None:
        next_row = 2

    thin   = Border(
                 left  =Side(style="thin", color="EEEEEE"),
                 right =Side(style="thin", color="EEEEEE"),
                 top   =Side(style="thin", color="EEEEEE"),
                 bottom=Side(style="thin", color="EEEEEE"),
             )
    today  = datetime.now().strftime("%d/%m/%Y")
    added, skipped = [], 0

    SPONS_COLOR = {"CONFIRMED": "C6F6D5", "LIKELY": "FEFCBF", "UNKNOWN": "FFF5F5"}

    for job in jobs:
        jid = _job_id(job["title"], job["company"], job["url"])
        if jid in existing_ids:
            skipped += 1
            continue

        even       = next_row % 2 == 0
        row_fill   = PatternFill("solid", fgColor="FFFFFF" if even else "F8FAFC")
        user_fill  = PatternFill("solid", fgColor="EBF8FF" if even else "E3F2FD")
        why_fill   = PatternFill("solid", fgColor="FFFBEB" if even else "FFF8E1")
        spons_col  = SPONS_COLOR.get(job.get("sponsorship_status", ""), "FFFFFF")
        spons_fill = PatternFill("solid", fgColor=spons_col)

        row_data = [
            jid,
            today,
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("salary", ""),
            job.get("source", ""),
            job.get("sponsorship_status", ""),
            job.get("_score", ""),
            job.get("url", ""),
            "",          # K Applied?
            "",          # L Date Applied
            "Pending",   # M Outcome
            "",          # N Why No?
            "",          # O Notes
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=9)
            cell.border    = thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

            if col_idx == 8:             # sponsorship — colour-coded
                cell.fill = spons_fill
            elif col_idx == 14:          # Why No — warm yellow
                cell.fill = why_fill
                cell.font = Font(name="Arial", size=9, italic=True, color="744210")
            elif col_idx >= 11:          # other user columns — light blue
                cell.fill = user_fill
            else:                        # bot columns — alternating white/grey
                cell.fill = row_fill

            if col_idx == 10 and value:  # URL — clickable hyperlink
                cell.hyperlink = value
                cell.value     = "Apply Here"
                cell.font      = Font(name="Arial", size=9, color="1A6FA8", underline="single")

        ws.row_dimensions[next_row].height = 16
        next_row += 1
        added.append(job)
        existing_ids.add(jid)

    try:
        wb.save(TRACKER_FILE)
        log.info("[TRACKER] Saved: %d added, %d already existed", len(added), skipped)
    except Exception as e:
        log.error("[TRACKER] Save failed: %s", e)

    return added, skipped
